import asyncio
import base64
import csv
import io
import json
import os
import re
 
import gspread
from aiogram import Bot, Dispatcher, types
from aiogram.filters import Command, CommandStart
from aiogram.types import KeyboardButton, ReplyKeyboardMarkup
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from openai import OpenAI
 
try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
 
# =========================================
# CLIENTS
# =========================================
BOT_TOKEN = os.getenv("BOT_TOKEN")
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")
 
bot = Bot(token=BOT_TOKEN)
dp = Dispatcher()
ai = OpenAI(api_key=OPENAI_API_KEY)
 
# =========================================
# GOOGLE
# =========================================
SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
]
creds = Credentials.from_service_account_file(
    "/etc/secrets/just-sunrise-501012-t4-829cee1f1963.json",
    scopes=SCOPES,
)
gs_client = gspread.authorize(creds)
sheets_svc = build("sheets", "v4", credentials=creds)
default_spreadsheet = gs_client.open("TEST BOT")
 
# =========================================
# STATE
# =========================================
user_modes: dict[int, str] = {}
chat_history: dict[int, list] = {}
sheet_history: dict[int, list] = {}
active_spreadsheets: dict[int, gspread.Spreadsheet] = {}
MAX_CHAT_HISTORY = 40
MAX_SHEET_HISTORY = 20
MAX_AGENT_STEPS = 20  # максимум шагов агента за одно сообщение
 
# =========================================
# KEYBOARD
# =========================================
main_keyboard = ReplyKeyboardMarkup(
    keyboard=[[KeyboardButton(text="💬 Chat"), KeyboardButton(text="📊 Sheets")]],
    resize_keyboard=True,
    input_field_placeholder="Choose mode",
)
 
# =========================================
# TOOLS — то что GPT может вызывать сам
# =========================================
TOOLS = [
    {
        "type": "function",
        "function": {
            "name": "read_range",
            "description": "Read cell values from a sheet. Use this BEFORE editing to understand current data.",
            "parameters": {
                "type": "object",
                "properties": {
                    "sheet_name": {"type": "string", "description": "Sheet tab name"},
                    "range": {"type": "string", "description": "A1 notation, e.g. 'A1:H30'"},
                },
                "required": ["sheet_name", "range"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "write_range",
            "description": "Write 2D array of values to a sheet starting from a cell.",
            "parameters": {
                "type": "object",
                "properties": {
                    "sheet_name": {"type": "string"},
                    "start_cell": {"type": "string", "description": "e.g. 'A1'"},
                    "values": {
                        "type": "array",
                        "items": {"type": "array", "items": {"type": "string"}},
                        "description": "2D array of values",
                    },
                },
                "required": ["sheet_name", "start_cell", "values"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "clear_range",
            "description": "Clear (empty) a range of cells without deleting the rows.",
            "parameters": {
                "type": "object",
                "properties": {
                    "sheet_name": {"type": "string"},
                    "range": {"type": "string", "description": "e.g. 'A20:Z26'"},
                },
                "required": ["sheet_name", "range"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "delete_rows",
            "description": "Delete entire rows from a sheet (shifts rows up).",
            "parameters": {
                "type": "object",
                "properties": {
                    "sheet_name": {"type": "string"},
                    "start_row": {"type": "integer", "description": "1-based row index"},
                    "end_row": {"type": "integer", "description": "1-based row index inclusive"},
                },
                "required": ["sheet_name", "start_row", "end_row"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "create_sheet",
            "description": "Create a new tab/sheet in the spreadsheet.",
            "parameters": {
                "type": "object",
                "properties": {
                    "title": {"type": "string"},
                    "rows": {"type": "integer", "default": 200},
                    "cols": {"type": "integer", "default": 30},
                },
                "required": ["title"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "delete_sheet",
            "description": "Delete a tab/sheet by name.",
            "parameters": {
                "type": "object",
                "properties": {
                    "sheet_name": {"type": "string"},
                },
                "required": ["sheet_name"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "list_sheets",
            "description": "List all sheet tabs in the spreadsheet.",
            "parameters": {"type": "object", "properties": {}},
        },
    },
    {
        "type": "function",
        "function": {
            "name": "format_header",
            "description": "Make a row bold with background color (for table headers).",
            "parameters": {
                "type": "object",
                "properties": {
                    "sheet_name": {"type": "string"},
                    "row": {"type": "integer", "description": "1-based row number"},
                    "num_cols": {"type": "integer", "description": "How many columns to format"},
                    "color": {
                        "type": "string",
                        "description": "Color preset: 'blue', 'green', 'dark', 'gray'",
                        "enum": ["blue", "green", "dark", "gray"],
                    },
                },
                "required": ["sheet_name", "row", "num_cols"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "done",
            "description": "Signal that the task is complete. Provide a short summary in Russian.",
            "parameters": {
                "type": "object",
                "properties": {
                    "summary": {"type": "string", "description": "Short summary in Russian of what was done"},
                    "url": {"type": "string", "description": "Optional direct link to the sheet"},
                },
                "required": ["summary"],
            },
        },
    },
]
 
# =========================================
# TOOL EXECUTORS
# =========================================
HEADER_COLORS = {
    "blue":  (0.18, 0.33, 0.55),
    "green": (0.13, 0.37, 0.20),
    "dark":  (0.15, 0.15, 0.20),
    "gray":  (0.45, 0.45, 0.50),
}
 
def _get_ws(spreadsheet: gspread.Spreadsheet, sheet_name: str) -> gspread.Worksheet:
    try:
        return spreadsheet.worksheet(sheet_name)
    except Exception:
        all_titles = [s.title for s in spreadsheet.worksheets()]
        match = next((t for t in all_titles if sheet_name.lower() in t.lower()), None)
        if match:
            return spreadsheet.worksheet(match)
        raise ValueError(f"Лист '{sheet_name}' не найден. Доступные: {all_titles}")
 
def tool_read_range(spreadsheet, args):
    ws = _get_ws(spreadsheet, args["sheet_name"])
    values = ws.get(args["range"])
    if not values:
        return {"values": [], "note": "Range is empty"}
    return {"sheet": args["sheet_name"], "range": args["range"], "values": values}
 
def tool_write_range(spreadsheet, args):
    ws = _get_ws(spreadsheet, args["sheet_name"])
    ws.update(args["start_cell"], args["values"])
    return {"ok": True, "written_rows": len(args["values"]), "start": args["start_cell"]}
 
def tool_clear_range(spreadsheet, args):
    ws = _get_ws(spreadsheet, args["sheet_name"])
    ws.batch_clear([args["range"]])
    return {"ok": True, "cleared": args["range"]}
 
def tool_delete_rows(spreadsheet, args):
    ws = _get_ws(spreadsheet, args["sheet_name"])
    sheet_id = ws.id
    start = args["start_row"] - 1  # 0-based
    end = args["end_row"]          # exclusive
    sheets_svc.spreadsheets().batchUpdate(
        spreadsheetId=spreadsheet.id,
        body={"requests": [{
            "deleteDimension": {
                "range": {
                    "sheetId": sheet_id,
                    "dimension": "ROWS",
                    "startIndex": start,
                    "endIndex": end,
                }
            }
        }]}
    ).execute()
    return {"ok": True, "deleted_rows": f"{args['start_row']}-{args['end_row']}"}
 
def tool_create_sheet(spreadsheet, args):
    title = args["title"]
    existing = [s.title for s in spreadsheet.worksheets()]
    if title in existing:
        return {"ok": False, "error": f"Sheet '{title}' already exists"}
    ws = spreadsheet.add_worksheet(
        title=title,
        rows=args.get("rows", 200),
        cols=args.get("cols", 30),
    )
    return {"ok": True, "created": title, "sheet_id": ws.id}
 
def tool_delete_sheet(spreadsheet, args):
    all_sheets = spreadsheet.worksheets()
    if len(all_sheets) <= 1:
        return {"ok": False, "error": "Cannot delete last sheet"}
    ws = _get_ws(spreadsheet, args["sheet_name"])
    spreadsheet.del_worksheet(ws)
    return {"ok": True, "deleted": args["sheet_name"]}
 
def tool_list_sheets(spreadsheet, args):
    sheets = spreadsheet.worksheets()
    return {"sheets": [{"title": s.title, "id": s.id, "rows": s.row_count, "cols": s.col_count} for s in sheets]}
 
def tool_format_header(spreadsheet, args):
    ws = _get_ws(spreadsheet, args["sheet_name"])
    sheet_id = ws.id
    row = args["row"] - 1  # 0-based
    num_cols = args["num_cols"]
    color_key = args.get("color", "blue")
    r, g, b = HEADER_COLORS.get(color_key, HEADER_COLORS["blue"])
    sheets_svc.spreadsheets().batchUpdate(
        spreadsheetId=spreadsheet.id,
        body={"requests": [
            {"repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": row,
                    "endRowIndex": row + 1,
                    "startColumnIndex": 0,
                    "endColumnIndex": num_cols,
                },
                "cell": {"userEnteredFormat": {
                    "textFormat": {"bold": True, "foregroundColor": {"red": 1, "green": 1, "blue": 1}},
                    "backgroundColor": {"red": r, "green": g, "blue": b},
                    "horizontalAlignment": "CENTER",
                }},
                "fields": "userEnteredFormat(textFormat,backgroundColor,horizontalAlignment)",
            }},
            {"autoResizeDimensions": {
                "dimensions": {
                    "sheetId": sheet_id,
                    "dimension": "COLUMNS",
                    "startIndex": 0,
                    "endIndex": num_cols,
                }
            }},
        ]}
    ).execute()
    return {"ok": True, "formatted_row": args["row"], "sheet": args["sheet_name"]}
 
TOOL_MAP = {
    "read_range": tool_read_range,
    "write_range": tool_write_range,
    "clear_range": tool_clear_range,
    "delete_rows": tool_delete_rows,
    "create_sheet": tool_create_sheet,
    "delete_sheet": tool_delete_sheet,
    "list_sheets": tool_list_sheets,
    "format_header": tool_format_header,
}
 
def execute_tool(name: str, args: dict, spreadsheet: gspread.Spreadsheet) -> str:
    if name not in TOOL_MAP:
        return json.dumps({"error": f"Unknown tool: {name}"})
    try:
        result = TOOL_MAP[name](spreadsheet, args)
        return json.dumps(result, ensure_ascii=False)
    except Exception as e:
        return json.dumps({"error": str(e)})
 
# =========================================
# SYSTEM PROMPT
# =========================================
def system_prompt(spreadsheet_title: str) -> str:
    return f"""You are a Google Sheets agent with direct API access to "{spreadsheet_title}".
 
You work in an agent loop: you can call tools multiple times until the task is complete.
Always read data before editing it — never guess cell values.
 
RULES:
- Call read_range only when you genuinely need to see existing data before editing
- If the user provides data in the message (CSV content, values) — use it directly, do NOT read the sheet first
- If the task is "write X to column Y in sheet Z" and you have the value — just write_range immediately
- After completing all actions call done() with a short Russian summary in Russian
- Communicate with the user in Russian
- Be efficient: minimum tool calls needed to complete the task
 
STEP EFFICIENCY GUIDE (use as few steps as possible):
- "разнеси данные из репорта в колонку" → write_range → done  (data already provided, no need to read first)
- "убери пробелы в строках 20-26" → read_range → write_range(fixed) → done
- "перенеси строки 27-33 в строку 20" → read_range(27:33) → write_range(A20) → clear_range(27:33) → done
- "построй таблицу" → create_sheet → write_range → format_header → done
- "удали все вкладки кроме вкладки 5" → list_sheets → delete_sheet × N → done
- "найди нужную колонку и запиши" → read_range(row1 only for headers) → write_range → done
"""
 
# =========================================
# AGENT LOOP
# =========================================
async def run_agent(
    user_message,           # str or list (for image)
    spreadsheet: gspread.Spreadsheet,
    chat_id: int,
    status_callback,        # async fn(text) — отправляет промежуточный статус
) -> str:
    """
    Крутит agent loop: GPT вызывает инструменты сам пока не вызовет done().
    Возвращает финальный summary.
    """
    history = sheet_history.get(chat_id, [])[-MAX_SHEET_HISTORY:]
 
    # Добавляем сообщение пользователя
    if isinstance(user_message, str):
        history.append({"role": "user", "content": user_message})
    else:
        history.append({"role": "user", "content": user_message})
 
    messages = [{"role": "system", "content": system_prompt(spreadsheet.title)}] + history
 
    for step in range(MAX_AGENT_STEPS):
        response = ai.chat.completions.create(
            model="gpt-4.1",
            messages=messages,
            tools=TOOLS,
            tool_choice="auto",
            temperature=0,
        )
 
        msg = response.choices[0].message
        messages.append(msg)
 
        # Нет tool calls — GPT хочет ответить текстом (не должно быть, но на всякий)
        if not msg.tool_calls:
            final = msg.content or "Готово."
            _save_history(chat_id, history, msg)
            return final
 
        # Обрабатываем все tool calls
        for tc in msg.tool_calls:
            fn_name = tc.function.name
            fn_args = json.loads(tc.function.arguments)
 
            # done() — завершение
            if fn_name == "done":
                summary = fn_args.get("summary", "Готово.")
                url = fn_args.get("url", "")
                _save_history(chat_id, history, msg)
                if url:
                    return f"✅ {summary}\n🔗 {url}"
                return f"✅ {summary}"
 
            # Статус пользователю
            status = _tool_status(fn_name, fn_args)
            await status_callback(status)
 
            # Выполняем инструмент
            result = execute_tool(fn_name, fn_args, spreadsheet)
 
            # Добавляем результат в messages
            messages.append({
                "role": "tool",
                "tool_call_id": tc.id,
                "content": result,
            })
 
    _save_history(chat_id, history, None)
    return "⚠️ Достигнут лимит шагов. Задача могла выполниться частично."
 
def _tool_status(fn_name: str, args: dict) -> str:
    labels = {
        "read_range": lambda a: f"📖 Читаю {a.get('sheet_name')}!{a.get('range')}...",
        "write_range": lambda a: f"✏️ Записываю в {a.get('sheet_name')}!{a.get('start_cell')}...",
        "clear_range": lambda a: f"🧹 Очищаю {a.get('sheet_name')}!{a.get('range')}...",
        "delete_rows": lambda a: f"🗑 Удаляю строки {a.get('start_row')}-{a.get('end_row')}...",
        "create_sheet": lambda a: f"📄 Создаю лист «{a.get('title')}»...",
        "delete_sheet": lambda a: f"🗑 Удаляю лист «{a.get('sheet_name')}»...",
        "list_sheets":  lambda a: "📋 Смотрю список листов...",
        "format_header": lambda a: f"🎨 Форматирую заголовок в {a.get('sheet_name')}...",
    }
    fn = labels.get(fn_name)
    return fn(args) if fn else f"⚙️ {fn_name}..."
 
def _save_history(chat_id, history, last_msg):
    if last_msg:
        history.append({"role": "assistant", "content": last_msg.content or ""})
    sheet_history[chat_id] = history[-MAX_SHEET_HISTORY:]
 
# =========================================
# HANDLERS
# =========================================
@dp.message(CommandStart())
async def start(message: types.Message):
    user_modes[message.chat.id] = "chat"
    active_spreadsheets[message.chat.id] = default_spreadsheet
    await message.answer("Choose mode:", reply_markup=main_keyboard)
 
@dp.message(Command("chat"))
async def cmd_chat(message: types.Message):
    user_modes[message.chat.id] = "chat"
    await message.answer("💬 Chat mode enabled", reply_markup=main_keyboard)
 
@dp.message(Command("sheet"))
async def cmd_sheet(message: types.Message):
    user_modes[message.chat.id] = "sheet"
    await message.answer("📊 Sheet mode enabled", reply_markup=main_keyboard)
 
@dp.message(lambda m: m.text == "💬 Chat")
async def btn_chat(message: types.Message):
    user_modes[message.chat.id] = "chat"
    await message.answer("💬 Chat mode enabled", reply_markup=main_keyboard)
 
@dp.message(lambda m: m.text == "📊 Sheets")
async def btn_sheet(message: types.Message):
    user_modes[message.chat.id] = "sheet"
    await message.answer("📊 Sheet mode enabled", reply_markup=main_keyboard)
 
# ---- PHOTO ----
@dp.message(lambda m: m.photo)
async def handle_photo(message: types.Message):
    try:
        chat_id = message.chat.id
        mode = user_modes.get(chat_id, "chat")
 
        photo = message.photo[-1]
        file = await bot.get_file(photo.file_id)
        b64 = base64.b64encode((await bot.download_file(file.file_path)).read()).decode()
 
        if mode == "chat":
            history = chat_history.get(chat_id, [])
            user_text = message.caption or "Analyze image"
            resp = ai.chat.completions.create(
                model="gpt-4.1",
                messages=[
                    {"role": "system", "content": "You are a helpful assistant. Speak naturally in Russian."}
                ] + history + [{"role": "user", "content": [
                    {"type": "text", "text": user_text},
                    {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{b64}"}},
                ]}],
                temperature=0.7,
            )
            answer = resp.choices[0].message.content
            history.append({"role": "user", "content": user_text})
            history.append({"role": "assistant", "content": answer})
            chat_history[chat_id] = history[-MAX_CHAT_HISTORY:]
            await message.answer(answer)
 
        elif mode == "sheet":
            spreadsheet = active_spreadsheets.get(chat_id, default_spreadsheet)
            caption = message.caption or "Reproduce this table structure exactly. Create a new sheet, write all headers and sample rows, format the header row."
            status_msg = await message.answer("🔍 Анализирую изображение...")
 
            sent_statuses = [status_msg]
            async def update_status(text):
                try:
                    await sent_statuses[-1].edit_text(text)
                except Exception:
                    m = await message.answer(text)
                    sent_statuses.append(m)
 
            user_content = [
                {"type": "text", "text": caption},
                {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{b64}", "detail": "high"}},
            ]
            result = await run_agent(user_content, spreadsheet, chat_id, update_status)
            await sent_statuses[-1].edit_text(result)
 
    except Exception as e:
        await message.answer(f"PHOTO ERROR:\n{e}")
 
# =========================================
# FILE PARSERS
# =========================================
def parse_csv(data: bytes) -> list[list[str]]:
    """Парсит CSV в список строк."""
    text = data.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    return [row for row in reader if any(cell.strip() for cell in row)]
 
def parse_xlsx(data: bytes) -> list[list[str]]:
    """Парсит XLSX в список строк."""
    if not HAS_OPENPYXL:
        raise RuntimeError("openpyxl не установлен")
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(values_only=True):
        str_row = [str(cell) if cell is not None else "" for cell in row]
        if any(c.strip() for c in str_row):
            rows.append(str_row)
    return rows
 
def rows_to_text(rows: list[list[str]], max_rows: int = 200) -> str:
    """Конвертирует строки в читаемый текст для GPT."""
    if not rows:
        return "(пустой файл)"
    total = len(rows)
    sample = rows[:max_rows]
    lines = ["\t".join(r) for r in sample]
    result = "\n".join(lines)
    if total > max_rows:
        result += f"\n... (показано {max_rows} из {total} строк)"
    return result
 
# ---- DOCUMENT (CSV / XLSX / другие файлы) ----
@dp.message(lambda m: m.document)
async def handle_document(message: types.Message):
    try:
        chat_id = message.chat.id
        mode = user_modes.get(chat_id, "chat")
        doc = message.document
        fname = (doc.file_name or "").lower()
        caption = message.caption or ""
 
        # Скачиваем файл
        file = await bot.get_file(doc.file_id)
        data = (await bot.download_file(file.file_path)).read()
 
        # CHAT MODE — просто анализируем текстом
        if mode == "chat":
            if fname.endswith(".csv"):
                rows = parse_csv(data)
                content = rows_to_text(rows)
                prompt = f"{caption}\n\nСодержимое файла {doc.file_name}:\n{content}" if caption else f"Вот содержимое файла {doc.file_name}:\n{content}"
            elif fname.endswith((".xlsx", ".xls")):
                rows = parse_xlsx(data)
                content = rows_to_text(rows)
                prompt = f"{caption}\n\nСодержимое файла {doc.file_name}:\n{content}" if caption else f"Вот содержимое файла {doc.file_name}:\n{content}"
            else:
                await message.answer("В режиме чата поддерживаю CSV и XLSX файлы.")
                return
 
            history = chat_history.get(chat_id, [])
            history.append({"role": "user", "content": prompt})
            resp = ai.chat.completions.create(
                model="gpt-4.1",
                messages=[
                    {"role": "system", "content": "You are a helpful AI assistant. Speak naturally in Russian."}
                ] + history[-MAX_CHAT_HISTORY:],
                temperature=0.7,
            )
            answer = resp.choices[0].message.content
            history.append({"role": "assistant", "content": answer})
            chat_history[chat_id] = history[-MAX_CHAT_HISTORY:]
            await message.answer(answer)
            return
 
        # SHEET MODE — передаём данные в agent loop
        if mode == "sheet":
            spreadsheet = active_spreadsheets.get(chat_id, default_spreadsheet)
 
            if fname.endswith(".csv"):
                rows = parse_csv(data)
            elif fname.endswith((".xlsx", ".xls")):
                rows = parse_xlsx(data)
            else:
                await message.answer(f"⚠️ Формат не поддерживается. Жди CSV или XLSX.")
                return
 
            content = rows_to_text(rows)
            row_count = len(rows)
            col_count = max((len(r) for r in rows), default=0)
 
            # Формируем промпт для агента
            user_prompt = (
                f"{caption}\n\n" if caption else ""
            ) + (
                f"Файл: {doc.file_name} ({row_count} строк, {col_count} колонок)\n\n"
                f"Данные:\n{content}\n\n"
                f"Задача: загрузи эти данные в таблицу «{spreadsheet.title}». "
                f"Создай новый лист с названием из первой строки или именем файла, "
                f"запиши все данные, отформатируй заголовок."
            )
 
            status_msg = await message.answer(f"📂 Читаю {doc.file_name} ({row_count} строк)...")
            sent_statuses = [status_msg]
 
            async def update_status(text_s):
                try:
                    await sent_statuses[-1].edit_text(text_s)
                except Exception:
                    m = await message.answer(text_s)
                    sent_statuses.append(m)
 
            result = await run_agent(user_prompt, spreadsheet, chat_id, update_status)
            try:
                await sent_statuses[-1].edit_text(result)
            except Exception:
                await message.answer(result)
 
    except Exception as e:
        await message.answer(f"FILE ERROR:\n{e}")
 
# ---- TEXT ----
@dp.message()
async def handle_text(message: types.Message):
    try:
        text = message.text
        if not text or text.startswith("/") or text in ["💬 Chat", "📊 Sheets"]:
            return
 
        chat_id = message.chat.id
        mode = user_modes.get(chat_id, "chat")
 
        # CHAT MODE
        if mode == "chat":
            history = chat_history.get(chat_id, [])
            history.append({"role": "user", "content": text})
            resp = ai.chat.completions.create(
                model="gpt-4.1",
                messages=[
                    {"role": "system", "content": "You are a helpful AI assistant. Speak naturally in Russian."}
                ] + history[-MAX_CHAT_HISTORY:],
                temperature=0.7,
            )
            answer = resp.choices[0].message.content
            history.append({"role": "assistant", "content": answer})
            chat_history[chat_id] = history[-MAX_CHAT_HISTORY:]
            await message.answer(answer)
            return
 
        # SHEET MODE
        if mode == "sheet":
            # Подключение нового документа по URL
            urls = re.findall(r'https://docs\.google\.com/spreadsheets/[^\s]+', text)
            if urls:
                try:
                    opened = gs_client.open_by_url(urls[0])
                    active_spreadsheets[chat_id] = opened
                    sheet_history[chat_id] = []
                    await message.answer(f'✅ Подключился к "{opened.title}"')
                except Exception as e:
                    await message.answer(f"❌ Не удалось открыть:\n{e}")
                return
 
            spreadsheet = active_spreadsheets.get(chat_id, default_spreadsheet)
 
            # Статус — редактируем одно сообщение
            status_msg = await message.answer("⚙️ Работаю...")
            sent_statuses = [status_msg]
 
            async def update_status(text_s):
                try:
                    await sent_statuses[-1].edit_text(text_s)
                except Exception:
                    m = await message.answer(text_s)
                    sent_statuses.append(m)
 
            result = await run_agent(text, spreadsheet, chat_id, update_status)
            try:
                await sent_statuses[-1].edit_text(result)
            except Exception:
                await message.answer(result)
 
    except Exception as e:
        await message.answer(f"ERROR:\n{e}")
 
# =========================================
# MAIN
# =========================================
async def main():
    await dp.start_polling(bot)
 
if __name__ == "__main__":
    asyncio.run(main())
